"""
工程量清单智能匹配定额工具
主程序入口

用法:
    python quota_matcher.py <输入文件路径> [-o <输出文件路径>]
    python quota_matcher.py "D:\工作\清单.xlsx"

示例:
    python quota_matcher.py "D:\工作\工程量清单.xlsx"
    python quota_matcher.py "D:\工作\清单.docx" -o "D:\工作\结果.xlsx"
"""

import os
import sys
import argparse
import json
from pathlib import Path
from datetime import datetime
from typing import List, Dict

# 添加当前目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from quota_loader import QuotaLoader
from file_parser import FileParser
from minimax_matcher import MiniMaxMatcher
from local_matcher import LocalMatcher
from quantity_extractor import QuantityExtractor


class QuotaMatcher:
    """定额匹配主类"""

    def __init__(self, use_local: bool = False, use_vector: bool = False):
        self.quota_loader = QuotaLoader()
        self.file_parser = FileParser()
        self.quota_data = None
        self.matcher = None
        self.use_local = use_local
        self.use_vector = use_vector
        self.vector_store = None
        self.quantity_extractor = QuantityExtractor()

    def load_quota_data(self):
        """加载定额数据"""
        print("正在加载定额数据...")
        self.quota_data = self.quota_loader.load()
        print(f"已加载 {len(self.quota_data)} 条定额")

        # 初始化向量存储（如果启用）
        if self.use_vector:
            try:
                from vector_store import VectorStore
                print("正在初始化向量存储...")
                self.vector_store = VectorStore()
                # 检查向量索引是否存在
                if not self.vector_store.has_index():
                    print("正在构建向量索引（首次运行需要几分钟）...")
                    api_key = os.environ.get("MINIMAX_API_KEY")
                    self.vector_store.build_index(self.quota_data, api_key)
                print("向量存储初始化完成")
            except Exception as e:
                print(f"向量存储初始化失败: {e}")
                self.vector_store = None

        # 初始化匹配器
        if self.use_local:
            print("使用本地关键词匹配...")
            self.matcher = LocalMatcher(self.quota_data)
        else:
            print("正在初始化AI匹配器...")
            api_key = os.environ.get("MINIMAX_API_KEY")
            if not api_key:
                raise ValueError("未设置MiniMax API密钥，请设置环境变量 MINIMAX_API_KEY")
            self.matcher = MiniMaxMatcher(self.quota_data, api_key, self.vector_store)

    def process(self, input_file: str, output_file: str = None) -> str:
        """
        处理工程量清单

        Args:
            input_file: 输入文件路径（Excel或Word）
            output_file: 输出文件路径（Excel），默认保存到输入文件所在目录

        Returns:
            str: 输出文件路径
        """
        # 确定输出路径
        if output_file is None:
            input_path = Path(input_file)
            output_file = input_path.parent / f"{input_path.stem}_matched.xlsx"

        output_path = Path(output_file)

        print(f"\n开始处理: {input_file}")
        print(f"输出文件: {output_path}")

        # 1. 解析输入文件
        print("\n[1/4] 正在解析文件...")
        parse_result = self.file_parser.parse(input_file)

        if len(parse_result.items) == 0:
            print("警告：未能解析到任何工程量项目")
            return None

        print(f"解析到 {len(parse_result.items)} 个工程量项目")

        if parse_result.warnings:
            print("解析警告:")
            for w in parse_result.warnings[:5]:  # 只显示前5条
                print(f"  - {w}")

        # 2. 加载定额数据（如果尚未加载）
        if self.quota_data is None:
            self.load_quota_data()

        # 2.5 工程量识别（新增）
        print("\n[2.5/5] 正在识别工程量...")
        for item in parse_result.items:
            # 设置默认source（如果来自附表会由StructuredTableExtractor处理）
            if "source" not in item:
                item["source"] = item.get("sheet", "")  # sheet包含"附表"则source为table

        enhanced_items = self.quantity_extractor.extract(parse_result.items)
        parse_result.items = enhanced_items
        print(f"工程量识别完成，{len(enhanced_items)} 项")

        # 3. AI匹配
        print("\n[3/5] 正在AI匹配定额（此过程需要联网）...")
        results = self.matcher.batch_match(parse_result.items)

        # 4. 生成输出Excel
        print("\n[4/5] 正在生成输出Excel...")
        self._write_output(results, output_path)

        # 5. 统计
        print("\n[5/5] 处理完成！")
        self._print_statistics(results)

        return str(output_path)

    def _write_output(self, results: List[Dict], output_path: Path):
        """写入输出Excel（参考北京易玖格式）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            print("错误：需要安装 openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "定额匹配结果"

        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # 表头（5列格式）
        headers = [
            ("定额编号", 12),
            ("项目名称", 40),
            ("单位", 8),
            ("工程量", 12),
            ("备注", 35)
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col)].width = width

        # 设置行高
        ws.row_dimensions[1].height = 25

        # 数据行
        for idx, result in enumerate(results, 1):
            row = idx + 1

            # 定额编号
            code_cell = ws.cell(row=row, column=1, value=result.get("code", ""))
            code_cell.alignment = center_align
            code_cell.border = thin_border
            # 高亮待确认的编号（黄色背景）
            if result.get("need_confirm"):
                code_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # 项目名称（原始清单名称）
            quota_name = result.get("name", "")
            # 如果待确认，在名称后面标注
            if result.get("need_confirm") and "（待人工确认）" not in quota_name:
                quota_name = f"{quota_name}（待人工确认）"
            name_cell = ws.cell(row=row, column=2, value=result.get("original_name", quota_name))
            name_cell.border = thin_border

            # 单位
            ws.cell(row=row, column=3, value=result.get("unit", "")).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            # 工程量（原始清单工程量）
            qty_cell = ws.cell(row=row, column=4, value=result.get("original_quantity", ""))
            qty_cell.alignment = center_align
            qty_cell.border = thin_border

            # 备注（包含匹配说明和单位转换信息）
            note = result.get("note", "")
            # 添加匹配方式说明
            confidence = result.get("confidence", "low")
            if confidence == "high":
                match_note = "精确匹配"
            elif confidence == "medium":
                match_note = "模糊匹配"
            else:
                match_note = "待人工确认"
            if note:
                note = f"{match_note}；{note}"
            else:
                note = match_note
            note_cell = ws.cell(row=row, column=5, value=note)
            note_cell.border = thin_border

            # 为所有单元格应用边框
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border

        # 添加统计行
        stats_row = len(results) + 3
        ws.cell(row=stats_row, column=1, value="统计信息:")
        ws.cell(row=stats_row, column=1).font = Font(bold=True)

        high_count = sum(1 for r in results if r.get("confidence") == "high")
        medium_count = sum(1 for r in results if r.get("confidence") == "medium")
        low_count = sum(1 for r in results if r.get("confidence") == "low")

        ws.cell(row=stats_row + 1, column=1, value=f"精确匹配: {high_count}项")
        ws.cell(row=stats_row + 2, column=1, value=f"模糊匹配: {medium_count}项")
        ws.cell(row=stats_row + 3, column=1, value=f"待人工确认: {low_count}项")
        ws.cell(row=stats_row + 4, column=1, value=f"合计: {len(results)}项")

        # 添加时间戳
        ws.cell(row=stats_row + 6, column=1,
                value=f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(output_path)

    def _print_statistics(self, results: List[Dict]):
        """打印统计信息"""
        total = len(results)
        matched = sum(1 for r in results if r.get("code") and r.get("code") != "待人工确定")
        uncertain = sum(1 for r in results if r.get("need_confirm"))
        unmatched = sum(1 for r in results if not r.get("code"))

        print(f"\n统计信息:")
        print(f"  总项目数: {total}")
        print(f"  已匹配: {matched} ({matched/total*100:.1f}%)")
        print(f"  待人工确认: {uncertain} ({uncertain/total*100:.1f}%)")
        print(f"  未匹配: {unmatched} ({unmatched/total*100:.1f}%)")


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description="工程量清单智能匹配定额工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python quota_matcher.py "D:\\工作\\清单.xlsx"          # 使用AI匹配
  python quota_matcher.py "D:\\工作\\清单.xlsx" --local  # 使用本地匹配
  python quota_matcher.py "D:\\工作\\清单.docx" -o "D:\\工作\\结果.xlsx"

注意:
  AI匹配需要设置环境变量 MINIMAX_API_KEY
  set MINIMAX_API_KEY=sk-cp-...
        """
    )

    parser.add_argument("input_file", help="输入文件路径（Excel或Word）")
    parser.add_argument("-o", "--output", help="输出文件路径（Excel），默认保存到输入文件所在目录")
    parser.add_argument("--local", action="store_true", help="使用本地关键词匹配，不调用AI")
    parser.add_argument("--vector", action="store_true", help="使用向量搜索预筛选（需要先构建索引）")

    args = parser.parse_args()

    # 检查输入文件
    if not Path(args.input_file).exists():
        print(f"错误: 输入文件不存在: {args.input_file}")
        sys.exit(1)

    # 如果不使用本地匹配，检查API密钥
    if not args.local and not os.environ.get("MINIMAX_API_KEY"):
        print("错误: 未设置 MINIMAX_API_KEY 环境变量")
        print("请先设置: set MINIMAX_API_KEY=sk-cp-...")
        print("或使用 --local 参数进行本地匹配")
        sys.exit(1)

    try:
        matcher = QuotaMatcher(use_local=args.local, use_vector=args.vector)
        output_file = matcher.process(args.input_file, args.output)

        if output_file:
            print(f"\n处理完成！结果已保存到:\n  {output_file}")
        else:
            print("\n处理失败，未生成输出文件")
            sys.exit(1)

    except Exception as e:
        print(f"\n错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
