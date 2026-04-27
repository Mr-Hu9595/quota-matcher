# -*- coding: utf-8 -*-
import sys
import os

# 使用openpyxl创建Excel文件
try:
    from openpyxl import Workbook
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "定额套取"

# 写入表头
headers = ["序号", "定额编号", "定额名称", "单位", "工程量表达式", "工程量", "匹配方式", "备注"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 写入数据
data = [
    [1, "4-10-60", "接地跨接线安装 构架接地", "处", 1417.2, 1417.2, "精确匹配", "BV-4mm²黄绿双色线 1417.2米"],
    [2, "4-13-179", "暗装接线盒安装", "个", 131, 131, "精确匹配", "三防接线盒 G1 IP65 WF1 131个"],
    [3, "4-13-181", "明装防爆接线盒安装", "个", 393, 393, "精确匹配", "防爆接线盒 G1 ExdIICT4 393个"],
    [4, "4-13-181", "明装防爆接线盒安装", "个", 610, 610, "待人工确认", "防爆活接头 G1-G1 ExdIICT4 610个(注:数据库无直接匹配定额)"],
    [5, "4-7-5+4-7-6", "一般铁构件制作+安装", "t", "槽钢94.5m+角钢42m", 1.048, "精确匹配", "热镀锌槽钢10#+角钢L40*4 按重量换算"],
    [6, "4-7-5+4-7-6", "一般铁构件制作+安装", "t", "钢管57.6m+钢板19块", 0.942, "精确匹配", "DN80热镀锌钢管立柱+200*200*8mm钢板"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 保存文件
output_path = "C:/Users/Administrator/Desktop/3签证单定额.xlsx"
wb.save(output_path)
print(f"Excel文件已保存到: {output_path}")