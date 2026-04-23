# -*- coding: utf-8 -*-
"""
工程量清单处理业务层
只管流程编排，不碰数据
"""

import os
from pathlib import Path
from datetime import datetime
from typing import List, Dict

from ..data.quota_db import QuotaDB
from ..data.vector_index import VectorIndex
from ..data.rule_db import RuleDB
from ..engine.base import EngineABC, MatchResult
from ..engine.hybrid_engine import HybridEngine
from ..utils.logging import get_logger, get_match_logger

logger = get_logger()
match_logger = get_match_logger()


class QuotaMatcherBusiness:
    """
    工程量清单处理业务

    流程：文件解析 → 工程量识别 → 匹配 → 输出Excel
    """

    def __init__(self,
                 engine: EngineABC = None,
                 quota_db: QuotaDB = None,
                 vector_index: VectorIndex = None,
                 rule_db: RuleDB = None):
        """
        初始化业务层

        Args:
            engine: 匹配引擎（默认 HybridEngine）
            quota_db: 定额数据库
            vector_index: 向量索引
            rule_db: 规则数据库
        """
        self.engine = engine or HybridEngine(
            rule_db=rule_db or RuleDB(),
            vector_index=vector_index or VectorIndex()
        )
        self.quota_db = quota_db or QuotaDB()
        self.vector_index = vector_index or VectorIndex()
        self.rule_db = rule_db or RuleDB()

    def process(self, input_file: str, output_file: str = None) -> str:
        """
        处理工程量清单

        Args:
            input_file: 输入文件路径
            output_file: 输出文件路径（可选）

        Returns:
            输出文件路径
        """
        logger.info("=" * 50)
        logger.info(f"开始处理工程量清单")
        logger.info(f"输入文件: {input_file}")

        # 确定输出路径
        if output_file is None:
            input_path = Path(input_file)
            output_file = str(input_path.parent / f"{input_path.stem}_matched.xlsx")

        output_path = Path(output_file)

        # 1. 解析文件
        logger.info("[1/5] 解析文件...")
        items = self._parse_file(input_file)
        if not items:
            logger.warning("未解析到任何工程量项目")
            return None
        logger.info(f"解析完成: {len(items)} 个项目")

        # 2. 提取工程量
        logger.info("[2/5] 提取工程量...")
        items = self._extract_quantities(items)
        logger.info(f"工程量提取完成: {len(items)} 项")

        # 2.5 计算辅料配比
        logger.info("[2.5/5] 计算辅料配比...")
        accessory_items = self._calculate_accessories(items)
        if accessory_items:
            items.extend(accessory_items)
            logger.info(f"辅料配比完成，新增 {len(accessory_items)} 项")

        # 3. 匹配
        logger.info("[3/5] AI/规则匹配...")
        results = self.engine.batch_match(items)
        logger.info(f"匹配完成: {len(results)} 个结果")

        # 4. 写Excel
        logger.info("[4/5] 生成输出Excel...")
        self._write_output(results, output_path)
        logger.info(f"Excel已保存: {output_path}")

        # 5. 统计
        logger.info("[5/5] 处理完成")
        self._print_statistics(results)

        logger.info("=" * 50)

        return str(output_path)

    def _parse_file(self, input_file: str) -> List[Dict]:
        """解析文件"""
        from ..file_parser import FileParser

        parser = FileParser()
        parse_result = parser.parse(input_file)

        if parse_result.warnings:
            for w in parse_result.warnings[:5]:
                logger.warning(f"解析警告: {w}")

        return parse_result.items

    def _extract_quantities(self, items: List[Dict]) -> List[Dict]:
        """提取工程量"""
        from ..quantity_extractor import QuantityExtractor

        extractor = QuantityExtractor()

        for item in items:
            if "source" not in item:
                item["source"] = item.get("sheet", "")

        return extractor.extract(items)

    def _calculate_accessories(self, items: List[Dict]) -> List[Dict]:
        """
        计算辅料配比

        电信：每个摄像头配置 1个防爆接线箱(300*300*200) + 3根防爆挠性管
        电气：每个防爆灯配置 2个电力电缆头 + 1个接线盒
        """
        new_items = []

        # 统计电信相关数量
        telecom_camera_qty = 0
        for item in items:
            name = item.get("name", "").lower()
            if any(kw in name for kw in ["防爆固定摄像机", "防爆摄像机", "摄像头", "camera"]):
                telecom_camera_qty += item.get("quantity", 0)

        if telecom_camera_qty > 0:
            new_items.append({
                "name": "防爆接线箱 300×300×200",
                "quantity": telecom_camera_qty,
                "unit": "个",
                "sheet": "辅料配比",
                "original_unit": "个"
            })
            new_items.append({
                "name": "防爆挠性管",
                "quantity": telecom_camera_qty * 3,
                "unit": "根",
                "sheet": "辅料配比",
                "original_unit": "根"
            })

        # 统计电气相关数量
        elec_light_qty = 0
        for item in items:
            name = item.get("name", "")
            if any(kw in name for kw in ["防爆弯灯", "防爆平台灯", "防爆灯", "LED"]):
                is_accessory = any(
                    name.startswith(kw) for kw in ["电力电缆头", "接线盒", "防爆接线箱", "防爆挠性管"]
                )
                if not is_accessory:
                    elec_light_qty += item.get("quantity", 0)

        if elec_light_qty > 0:
            new_items.append({
                "name": "电力电缆头",
                "quantity": elec_light_qty * 2,
                "unit": "个",
                "sheet": "辅料配比",
                "original_unit": "个"
            })
            new_items.append({
                "name": "接线盒",
                "quantity": elec_light_qty,
                "unit": "个",
                "sheet": "辅料配比",
                "original_unit": "个"
            })

        return new_items

    def _write_output(self, results: List[Dict], output_path: Path):
        """写Excel输出"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            logger.error("需要安装 openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "定额匹配结果"

        # 样式定义
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = [
            ("定额编号", 12),
            ("项目名称", 40),
            ("单位", 8),
            ("工程量", 12),
            ("备注", 35)
        ]

        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col)].width = width

        ws.row_dimensions[1].height = 25

        # 数据行
        for idx, result in enumerate(results, 1):
            row = idx + 1

            # 定额编号
            code_cell = ws.cell(row=row, column=1, value=result.get("code", ""))
            code_cell.alignment = center_align
            code_cell.border = thin_border
            if result.get("need_confirm"):
                code_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # 项目名称
            quota_name = result.get("name", "")
            if result.get("need_confirm") and "（待人工确认）" not in quota_name:
                quota_name = f"{quota_name}（待人工确认）"
            name_cell = ws.cell(row=row, column=2, value=result.get("original_name", quota_name))
            name_cell.border = thin_border

            # 单位
            ws.cell(row=row, column=3, value=result.get("unit", "")).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            # 工程量
            qty_cell = ws.cell(row=row, column=4, value=result.get("original_quantity", ""))
            qty_cell.alignment = center_align
            qty_cell.border = thin_border

            # 备注
            confidence = result.get("confidence", "low")
            if confidence == "high":
                match_note = "精确匹配"
            elif confidence == "medium":
                match_note = "模糊匹配"
            else:
                match_note = "待人工确认"

            note = result.get("note", "")
            if note:
                note = f"{match_note}；{note}"
            else:
                note = match_note

            note_cell = ws.cell(row=row, column=5, value=note)
            note_cell.border = thin_border

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border

        # 统计行
        stats_row = len(results) + 3
        ws.cell(row=stats_row, column=1, value="统计信息:")
        ws.cell(row=stats_row, column=1).font = Font(bold=True)

        high_count = sum(1 for r in results if r.get("confidence") == "high")
        medium_count = sum(1 for r in results if r.get("confidence") == "medium")
        low_count = sum(1 for r in results if r.get("confidence") == "low")

        ws.cell(row=stats_row + 1, column=1, value=f"精确匹配: {high_count}项")
        ws.cell(row=stats_row + 2, column=1, value=f"模糊匹配: {medium_count}项")
        ws.cell(row=stats_row + 3, column=1, value=f"待人工确认: {low_count}项")
        ws.cell(row=stats_row + 4, column=1, value=f"合计: {len(results)}项")
        ws.cell(row=stats_row + 6, column=1, value=f"处理时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        ws.freeze_panes = "A2"
        wb.save(output_path)

    def _print_statistics(self, results: List[Dict]):
        """打印统计信息"""
        total = len(results)
        matched = sum(1 for r in results if r.get("code") and r.get("code") != "待人工确定")
        uncertain = sum(1 for r in results if r.get("need_confirm"))

        logger.info(f"统计信息:")
        logger.info(f"  总项目数: {total}")
        logger.info(f"  已匹配: {matched} ({matched/total*100:.1f}%)")
        logger.info(f"  待人工确认: {uncertain} ({uncertain/total*100:.1f}%)")


# 便捷函数
def process_workflow(input_file: str, output_file: str = None) -> str:
    """
    处理工程量清单的便捷函数

    Args:
        input_file: 输入文件路径
        output_file: 输出文件路径（可选）

    Returns:
        输出文件路径
    """
    business = QuotaMatcherBusiness()
    return business.process(input_file, output_file)
