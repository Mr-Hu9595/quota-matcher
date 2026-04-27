# -*- coding: utf-8 -*-
import sys
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "签证单定额"

# 写入表头
headers = ["序号", "定额编号", "工作名称", "单位", "工程量", "主材单价", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 签证单工程量清单及定额匹配
# 1. BV-4mm²接地跨接线 1417.2米 → 4-10-57 接地跨接线 m
# 2. 三防接线盒 131个 → 4-13-182 接线盒安装 个
# 3. 防爆接线盒 393个 → 4-13-181 防爆接线盒安装 个
# 4. 防爆活接头 610个 → 4-13-181（参考防爆接线盒） 个
# 5. 热镀锌槽钢10# 94.5米 → 4-8-20 槽钢安装 10m/根 → 9.45(10m)
# 6. 热镀锌角钢L40*4 42米 → 4-8-10 角钢安装 10m/根 → 4.2(10m)
# 7. 热镀锌钢管DN80 57.6米 → 8-1-441 钢管安装 10m → 5.76(10m)
# 8. 摄像头立柱（DN80+钢板）→ 4-7-5制作+4-7-6安装 t

# 主材市场价（2026年4月参考）
# BV-4mm²接地线：约4元/m
# 防爆接线盒：约50元/个
# 防爆活接头：约15元/个
# 热镀锌槽钢10#：约3700元/吨
# 热镀锌角钢L40*4：约3700元/吨
# 热镀锌钢管DN80：约4200元/吨
# Q235B钢板8mm：约3650元/吨

data = [
    [1, "4-10-57", "接地跨接线安装", "m", 1417.2, 4, "BV-4mm²黄绿双色线1417.2米"],
    [2, "4-13-182", "三防接线盒安装", "个", 131, 50, "三防接线盒 G1 IP65 WF1 131个"],
    [3, "4-13-181", "防爆接线盒安装", "个", 393, 50, "防爆接线盒 G1 ExdIICT4 393个"],
    [4, "4-13-181", "防爆活接头安装", "个", 610, 15, "防爆活接头 G1\"-G1\" 610个"],
    [5, "4-8-20", "热镀锌槽钢10#安装", "10m", 9.45, 3700, "槽钢94.5米→9.45×10m 单位换算:米→10m"],
    [6, "4-8-10", "热镀锌角钢L40*4安装", "10m", 4.2, 3700, "角钢42米→4.2×10m 单位换算:米→10m"],
    [7, "8-1-441", "热镀锌钢管DN80安装", "10m", 5.76, 4200, "DN80钢管57.6米→5.76×10m 单位换算:米→10m"],
    [8, "4-7-5", "一般铁构件制作", "t", 2.72, 3650, "立柱:DN80钢管2.247t+钢板0.477t=2.724t"],
    [9, "4-7-6", "一般铁构件安装", "t", 2.72, 0, "一般铁构件安装 仅含人工机械费"],
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 1:
            cell.alignment = Alignment(horizontal='center')
        elif col_idx in [4, 5]:
            cell.alignment = Alignment(horizontal='right')

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 40

# 保存文件
output_path = "C:/Users/Administrator/Desktop/3签证单定额.xlsx"
wb.save(output_path)
print(f"Excel文件已保存到: {output_path}")
